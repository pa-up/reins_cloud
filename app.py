"""
日時指定のための各種設定を本WEBアプリで編集
    日時指定のスクレイピング実行はLambda・S3とRender上
"""

from flask import Flask, render_template, request , send_file , redirect , url_for , session
import os
import sys
import csv
import openpyxl
import boto3
import time
import re
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select



# flaskアプリの明示
templates_path = "templates/"
static_path = "static/"
app = Flask(__name__ , template_folder=templates_path, static_folder=static_path)

# パスの定義
mail_excel_path = static_path + "/input_excel/email_pw.xlsx"
search_method_csv_path = static_path + "/csv/search_method.csv"
output_reins_csv_path_from_static = "/csv/output_reins.csv"
output_reins_csv_path = static_path + output_reins_csv_path_from_static
output_reins_excel_path = static_path + "/output_excel/output_reins.xlsx"


# 環境変数の取得
user_id , password = os.environ.get('SECRET_USER_ID') , os.environ.get('SECRET_PASSWORD')
s3_accesskey , s3_secretkey = os.environ.get('S3_ACCESSKEY') , os.environ.get('S3_SECRETKEY')
s3_bucket_name = os.environ.get('S3_BUCKET_NAME')


# s3の定義
s3_region = "ap-northeast-1"   # 東京(アジアパシフィック)：ap-northeast-1


def browser_setup(browse_visually = "no"):
    """ブラウザを起動する関数"""
    #ブラウザの設定
    options = webdriver.ChromeOptions()
    if browse_visually == "no":
        options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    #ブラウザの起動
    browser = webdriver.Chrome(options=options , service=ChromeService(ChromeDriverManager().install()))
    browser.implicitly_wait(3)
    return browser


def list_to_csv(to_csv_list: list , csv_path: str = "output.csv"):
    """ 多次元リストのデータをcsvファイルに保存する関数 """
    with open(csv_path, 'w' , encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(to_csv_list)


def html_table_tag_to_csv_list(table_tag_str: str, header_exist: bool = True):
    table_soup = BeautifulSoup(table_tag_str, 'html.parser')
    rows = []
    if header_exist:
        for tr in table_soup.find_all('tr'):
            cols = [] 
            for td in tr.find_all(['td', 'th']):
                cols.append(td.text.strip())
            rows.append(cols)
    else:
        for tbody in table_soup.find_all('tbody'):
            for tr in tbody.find_all('tr'):
                cols = [td.text.strip() for td in tr.find_all(['td', 'th'])]
                rows.append(cols)
    return rows


def get_building_number(page_count_info: str):
    # 正規表現を使用して物件の件数を抽出する関数
    print(f"page_count_info : {page_count_info}")
    numbers = re.findall(r'\d+', page_count_info)
    print(f"numbers : {numbers}")
    # 数字が3つ以上見つかった場合、それぞれの数字を返す
    if len(numbers) >= 3:
        start_number = int(numbers[0])
        end_number = int(numbers[1])
        total_number = int(numbers[2])
        return start_number, end_number, total_number
    else:
        return None
    

def normalize_string(input_string):
    """ スペース、特殊文字、連続した空白を削除し、文字列を小文字に変換し、正規化する関数 """
    normalized = re.sub(r'\s+', '', input_string).lower()
    return normalized




class Reins_Scraper:
    def __init__(self, driver: WebDriverWait):
        self.driver = driver
        self.wait_driver = WebDriverWait(driver, 10)
    
    def login_reins(self, user_id: str , password: str ,):
        login_flag = "No"
        # ログインボタンをクリック
        try:
            login_button = self.wait_driver.until(EC.element_to_be_clickable((By.ID, "login-button")))
            login_button.click()
        except:
            maintenance = self.wait_driver.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.p-card"))).text
            return maintenance

        # フォームにログイン認証情報を入力
        user_id_form = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text']")))
        user_id_form.send_keys(user_id)
        password_form = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password']")))
        password_form.send_keys(password)
        rule_element = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//input[@type='checkbox' and contains(following-sibling::label, 'ガイドライン')]")))
        rule_checkbox_form = rule_element.find_element(By.XPATH, "./following-sibling::label")
        rule_checkbox_form.click()
        time.sleep(0.5)
        login_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), 'ログイン')]")))
        login_button.click()
        login_flag = "OK"
        return login_flag

    def get_solding_or_rental_option(self):
        # ボタン「売買 物件検索」をクリック
        sold_building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '売買') and contains(text(), '物件検索')]")))
        sold_building_search_button.click()
        time.sleep(1)
        # 検索条件を取得
        display_search_method_link = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "(//div[@class='card p-card'])[1]"))).find_element(By.XPATH, ".//a[contains(span, '検索条件を表示')]")
        display_search_method_link.click()
        time.sleep(1)
        # 検索条件のリストを取得
        select_element = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-selectbox']//select")))
        search_method_element_list = select_element.find_elements(By.TAG_NAME, "option")
        solding_search_method_list = []
        for search_method_element in search_method_element_list:
            solding_search_method_list.append( search_method_element.text )
        # 前のページに戻る
        self.driver.back()

        # ボタン「売買 物件検索」をクリック
        rental_building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '賃貸') and contains(text(), '物件検索')]")))
        rental_building_search_button.click()
        time.sleep(1)
        # 検索条件を取得
        display_search_method_link = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "(//div[@class='card p-card'])[1]"))).find_element(By.XPATH, ".//a[contains(span, '検索条件を表示')]")
        display_search_method_link.click()
        time.sleep(1)
        # 検索条件のリストを取得
        select_element = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-selectbox']//select")))
        search_method_element_list = select_element.find_elements(By.TAG_NAME, "option")
        rental_search_method_list = []
        for search_method_element in search_method_element_list:
            rental_search_method_list.append( search_method_element.text )
        # 前のページに戻る
        self.driver.back()
        time.sleep(2)
        return solding_search_method_list , rental_search_method_list
        
    def scraping_solding_list(self , search_method_value: str , index_of_search_requirement: int):
        # 選択された検索方法をクリック
        if search_method_value == "search_solding":
            building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '売買') and contains(text(), '物件検索')]")))
            building_search_button.click()
            time.sleep(1)
        else:
            building_search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '賃貸') and contains(text(), '物件検索')]")))
            building_search_button.click()
            time.sleep(1)

        # 売買検索条件を選択
        display_search_method_link = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "(//div[@class='card p-card'])[1]"))).find_element(By.XPATH, ".//a[contains(span, '検索条件を表示')]")
        display_search_method_link.click()
        time.sleep(1)
        choice_search_method = Select(self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-selectbox']//select"))))
        choice_search_method.select_by_index(index_of_search_requirement)
        get_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '読込')]")))
        get_button.click()
        time.sleep(1)
        time.sleep(0.5)

        # 検索条件が存在するか判定
        exist_search_requirement_sentence = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[class*="modal"]'))).text
        if "エラー" in exist_search_requirement_sentence:
            to_csv_list = False
            self.driver.quit()
            return to_csv_list
        
        ok_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), 'OK')]")))
        ok_button.click()
        time.sleep(1)

        # 検索条件に基づいて検索実行
        search_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//div[@class='p-frame-bottom']//button[contains(text(), '検索')]")))
        search_button.click()
        time.sleep(2)

        # 物件リストが何ページあるかを判定
        time.sleep(2)
        page_count_info = self.wait_driver.until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.card-header"))).text
        match = re.search(r'(\d+)件', page_count_info)
        total_number = int( match.group(1) )
        left_page_count = total_number / 50 

        # リストを取得
        loop_count = 0
        all_list = []
        while True:
            # 印刷表示ボタンをクリック
            print_button = self.wait_driver.until(EC.presence_of_element_located((By.XPATH, "//button[contains(text(), '印刷')]")))
            print_button.click()
            time.sleep(2)
            
            # 現在のページのHTML要素を取得
            table_tag_str = self.wait_driver.until(EC.presence_of_element_located((By.TAG_NAME, "table"))).get_attribute('outerHTML')
            # tableタグの要素を多次元リストに変換
            if loop_count == 0:
                header_exist = True
            else:
                header_exist = False
            loop_count += 1

            to_csv_list = html_table_tag_to_csv_list(
                table_tag_str = table_tag_str , header_exist = header_exist ,
            )
            all_list.append(to_csv_list)

            if left_page_count >= 1:
                left_page_count -= 1
                # リストの表示ページへ戻る
                back_button = self.wait_driver.until(EC.element_to_be_clickable((By.CLASS_NAME, 'p-frame-backer')))
                back_button.click()
                time.sleep(2)
                # 次のリストを表示させるボタンをクリック
                next_list_button = self.wait_driver.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'li.page-item > button > span.p-pagination-next-icon')))
                next_list_button.click()
                time.sleep(2)

            else:
                break

        self.driver.quit()
        
        # 全ての多次元リストを連結
        to_csv_list = []
        for loop in range( len(all_list) ):
            to_csv_list.extend( all_list[loop] )    
        
        return to_csv_list
    

def csv_to_list(csv_path: str = "output.csv"):
    """ 多次元データを含むcsvからリストに変換 """
    data_list = []
    with open(csv_path, 'r' , encoding="utf-8-sig") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data_list.append(row)
    return data_list


def list_to_excel(to_excel_list: list , output_excel_path: str = "output.xlsx"):
    workbook = openpyxl.Workbook()
    print("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    print("ワークブックのアクティブ化完了 : sheet = workbook.active")

    # 多次元リストのサイズを取得(行ごとで列数に違いがあることを考慮)
    row_num , col_num = len(to_excel_list) , 0
    for row in range(row_num):
        predict_col = len(to_excel_list[row])
        if predict_col > col_num:
            col_num = predict_col
    print(f"row_num : {row_num}")
    print(f"col_num : {col_num}")

    for row in range(row_num):
        for col in range(col_num):
            try:
                print(f"pressed_cell_value : {to_excel_list[row][col]}")
                print(f"row , col : {row} , {col} \n \n")
                sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
            except IndexError:
                pass
    print("セルの編集可能が証明 : sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]")
    workbook.save(output_excel_path)
    

def get_search_option_from_csv(input_csv_path):
    """ 定期実行ツールがcsvファイルから検索方法と条件を取得する関数 """
    search_option_list = csv_to_list(input_csv_path)
    search_method_value = search_option_list[1][0]
    search_requirement = int( search_option_list[1][1] )
    return search_method_value , search_requirement


def update_search_csv_file(csv_path , search_method , search_requirement_index):
    """ 検索方法と条件のcsvファイルを編集する関数 """
    # CSVファイルを読み込みモードで開く
    with open(csv_path, 'r', newline='', encoding='utf-8') as file:
        # CSVファイルを読み込む
        csv_reader = csv.reader(file)
        rows = list(csv_reader)
    # 2行1列目と2行2列目を書き換える
    rows[1][0] = search_method
    rows[1][1] = str(search_requirement_index)
    # CSVファイルを書き込みモードで開く
    with open(csv_path, 'w', newline='', encoding='utf-8') as file:
        # 書き込み用のCSVライターを作成
        csv_writer = csv.writer(file)
        # 書き換えたデータをCSVファイルに書き込む
        csv_writer.writerows(rows)


def mail_list_from_excel(mail_excel_path):
    """ Excelファイルからメールのリスト(cc含)を取得する関数 """
    mail_list = []
    workbook = openpyxl.load_workbook(mail_excel_path)
    sheet = workbook.active
    receive_email_number = 100
    for index in range(receive_email_number):
        mail = sheet.cell(row = index + 2 , column = 3).value
        # emailかどうかを判定（「@」「.」の有無）
        if mail is not None:
            if '@' in mail and '.' in mail:
                mail_list.append(mail)
        else:
            break
    # ccのメールのリストを取得
    cc_mail_list = []
    for index in range(len(mail_list)):
        # D列以降を判定
        cc_mail_row_list = []
        for col in range(receive_email_number):
            cc_mail = sheet.cell(row = index + 2 , column = 4 + col).value
            # emailかどうかを判定（「@」「.」の有無）
            if cc_mail is not None:
                if '@' in cc_mail and '.' in cc_mail:
                    cc_mail_row_list.append(cc_mail)
            else:
                break
        cc_mail_list.append(cc_mail_row_list)
    # 送信元メールアドレスとアプリパスワードを取得
    from_email = sheet.cell(row = 2 , column = 1).value
    from_email_smtp_password = sheet.cell(row = 2 , column = 2).value
    return mail_list , cc_mail_list , from_email , from_email_smtp_password


def mail_list_to_excel(receive_mail_list , mail_excel_path):
    """ メールリストのエクセルファイルを変更する関数 """
    workbook = openpyxl.load_workbook(mail_excel_path)
    sheet = workbook.active
    
    # C列の2行目から順にメールリストの要素を書き込む
    start_row = 2  # 2行目から始める例
    for receive_mail in receive_mail_list:
        if '@' in receive_mail and '.' in receive_mail:
            sheet.cell(row = start_row, column = 3, value = receive_mail)
        start_row += 1
    print(f"len(receive_mail_list) : {len(receive_mail_list)}")
    # メールアドレスが削除されたケース
    loop_count = len(receive_mail_list)
    end_flag = False
    while end_flag == False:
        delete_cell = sheet.cell(row = 2 + loop_count , column = 3).value
        print(f"delete_cell : {delete_cell}")
        if delete_cell != None:
            sheet.cell(row = 2 + loop_count , column = 3).value = None
            print(f"delete_cell : {sheet.cell(row = 2 + loop_count , column = 3).value}")
            loop_count += 1
            print("\n")
        else:
            break
    # 変更を保存
    workbook.save(mail_excel_path)



class ManipulateS3:
    def __init__(self , accesskey , secretkey , bucket_name , region = "ap-northeast-1"):
        self.region = region  # 東京(アジアパシフィック)：ap-northeast-1
        self.accesskey = accesskey
        self.secretkey = secretkey
        self.bucket_name = bucket_name
        self.s3 = boto3.client('s3', aws_access_key_id=self.accesskey, aws_secret_access_key=self.secretkey, region_name=self.region)
    
    def get_file_name_from_file_path(self , file_path):
        """ パスからファイル名のみを抽出する関数
            s3にフォルダを作成し、ファイルをアップロードする場合は、この関数を使わずに、file_pathにフォルダ名を含める
        """
        file_name_from_path = file_path[file_path.rfind('/') + 1 : ]  # ファイルパスからファイル名のみを抽出
        return file_name_from_path
    
    def s3_file_upload(self , file_path):
        """ s3の特定のバケットにファイルをアップロードし、そのファイルのURLも取得する関数
            s3上のファイルが同一のファイル名であれば、s3内で上書き保存される
        """
        key_name = self.get_file_name_from_file_path(file_path)
        # s3へファイルをアップロード
        self.s3.upload_file(file_path, self.bucket_name, key_name)
        # S3へアップロードしたファイルへのURLを取得する
        s3_url = self.s3.generate_presigned_url(
            ClientMethod='get_object',
            Params={'Bucket': self.bucket_name, 'Key': key_name},
            ExpiresIn=3600,
            HttpMethod='GET'
        )
        return s3_url
    
    def s3_file_download(self , local_upload_path):
        """ s3の特定のバケットからファイル名で検索し、一致するファイルをダウンロードする関数
            local_file_pathのファイル名はs3で取得予定のファイル名を同一にする
        """
        key_name = self.get_file_name_from_file_path(local_upload_path)
        print(f"key_name : {key_name}")
        self.s3.download_file(self.bucket_name, key_name, local_upload_path)


# S3からメール情報や検索条件を取得し、静的フォルダに格納
manipulate_s3 = ManipulateS3(
    region = s3_region ,
    accesskey = s3_accesskey ,
    secretkey = s3_secretkey ,
    bucket_name = s3_bucket_name ,
)



@app.route('/')
def index():
        return render_template("index.html")


@app.route('/order_scraping')
def order_scraping():
    global reins_sraper , solding_search_method_list , rental_search_method_list

    # ページにアクセス
    searched_url = "https://system.reins.jp/"
    driver = browser_setup()
    reins_sraper = Reins_Scraper(driver)
    driver.get(searched_url)

    # ログイン突破
    user_id , password = "127128210603" , "euc3ta"
    login_flag = reins_sraper.login_reins(user_id , password)
    if login_flag == "OK":
        solding_search_method_list , rental_search_method_list = reins_sraper.get_solding_or_rental_option()
        print(f"solding_search_method_list : {solding_search_method_list} \n")
        print(f"rental_search_method_list : {rental_search_method_list} \n")

        return render_template(
            "order_scraping.html" ,
            solding_search_method_list = solding_search_method_list ,
            rental_search_method_list = rental_search_method_list ,
            login_flag = login_flag ,
        )
    else:
        return render_template(
            "order_scraping.html" ,
            login_flag = login_flag ,
        )


@app.route('/result', methods=['GET', 'POST'])
def result():
    global reins_sraper , solding_search_method_list , rental_search_method_list
    if request.method == 'POST' and request.form['start_scraping'] == "true":
        # フォームから検索方法を取得
        search_method_value = request.form['search_method_value']
        # フォームから検索条件を取得
        search_requirement = request.form['solding'] if search_method_value == 'search_solding' else request.form['rental']
        print(f"search_requirement : {search_requirement} \n")

        if search_method_value == "search_solding":
            index_of_search_requirement = solding_search_method_list.index(search_requirement)
            print(f"index_of_search_requirement : {index_of_search_requirement} \n")
        else:
            index_of_search_requirement = rental_search_method_list.index(search_requirement)
            print(f"index_of_search_requirement : {index_of_search_requirement} \n")

        
        # リストの取得実行
        to_csv_list = reins_sraper.scraping_solding_list(search_method_value , index_of_search_requirement)
        if to_csv_list == False:
            return render_template(
                "result.html" ,
                search_method_value = search_method_value ,
                search_requirement = search_requirement ,
                no_exist_search_requirement = True ,
            )
        
        # リストをCSVファイルに保存
        list_to_csv(to_csv_list = to_csv_list , csv_path = output_reins_csv_path ,)
        # リストをExcelファイルに保存
        list_to_excel(to_csv_list , output_reins_excel_path)

        return render_template(
            "result.html" ,
            search_method_value = search_method_value ,
            search_requirement = search_requirement ,
            csv_path_from_static = output_reins_csv_path_from_static ,
        )
    

@app.route('/schedule_search')
def schedule_search():
    global reins_sraper , solding_search_method_list , rental_search_method_list

    # ページにアクセス
    searched_url = "https://system.reins.jp/"
    driver = browser_setup()
    reins_sraper = Reins_Scraper(driver)
    driver.get(searched_url)

    # ログイン突破
    user_id , password = "127128210603" , "euc3ta"
    login_flag = reins_sraper.login_reins(user_id , password)
    if login_flag == "OK":
        # 現時点での全ての検索条件の選択肢を表示
        solding_search_method_list , rental_search_method_list = reins_sraper.get_solding_or_rental_option()
        print(f"solding_search_method_list : {solding_search_method_list} \n")
        print(f"rental_search_method_list : {rental_search_method_list} \n")

        # 検索条件をcsvファイルをS3から取得
        manipulate_s3.s3_file_download(search_method_csv_path)
        time.sleep(2)

        # 日時実行の検索条件をcsvファイルから取得
        search_method_value , index_of_search_requirement = get_search_option_from_csv(search_method_csv_path)
        # 日時実行の検索方法と検索条件を文字列で取得
        if search_method_value == "search_solding":
            schedule_search_method = "売買検索"
            schedule_search_requirement = solding_search_method_list[index_of_search_requirement]
        else:
            schedule_search_method = "賃貸検索"
            schedule_search_requirement = rental_search_method_list[index_of_search_requirement]

        return render_template(
            "schedule_search.html" ,
            schedule_search_method = schedule_search_method ,
            schedule_search_requirement = schedule_search_requirement ,
            solding_search_method_list = solding_search_method_list ,
            rental_search_method_list = rental_search_method_list ,
            login_flag = login_flag ,
        )
    else:
        return render_template(
            "schedule_search.html" ,
            login_flag = login_flag ,
        )


@app.route('/search_result', methods=['GET', 'POST'])
def search_result():
    global reins_sraper , solding_search_method_list , rental_search_method_list

    if request.method == 'POST':
        # フォームから検索方法を取得
        search_method_value = request.form['search_method_value']
        # フォームから検索条件を取得
        search_requirement = request.form['solding'] if search_method_value == 'search_solding' else request.form['rental']
        print(f"search_requirement : {search_requirement} \n")

        # 検索条件の番号を変更
        if search_method_value == "search_solding":
            index_of_search_requirement = solding_search_method_list.index(search_requirement)
        else:
            index_of_search_requirement = rental_search_method_list.index(search_requirement)
        
        # csvファイルの検索方法と条件を書き換える
        update_search_csv_file(search_method_csv_path , search_method_value , index_of_search_requirement)
        # S3上のcsvファイルを更新
        manipulate_s3.s3_file_upload(search_method_csv_path)

        return render_template(
            "result_search.html" ,
            search_method_value = search_method_value ,
            search_requirement = search_requirement ,
        )
    


@app.route('/schedule_mail', methods=['GET', 'POST'])
def schedule_mail():
    global receive_mail_list
    if request.method == 'POST':
        # フォームから変更されたメールのリストを取得
        receive_mail_list_str = request.form['mail_list']
        receive_mail_list = re.split(r'\s+', receive_mail_list_str.strip())
        print(f"変更後 : ")
        print(receive_mail_list)
        # Excelファイルに変更
        mail_list_to_excel(receive_mail_list , mail_excel_path)
        time.sleep(2)
        # S3上のExcelファイルを更新
        manipulate_s3.s3_file_upload(mail_excel_path)
        time.sleep(2)

        # 再びページを表示
        receive_mail_list , cc_mail_list , from_email , \
            from_email_smtp_password = mail_list_from_excel(mail_excel_path)
        print(f"更新後 : ")
        print(receive_mail_list)
        receive_mail_list_str = "\n".join(receive_mail_list)
        return render_template(
            "schedule_mail.html" ,
            receive_mail_list_str = receive_mail_list_str ,
            from_email = from_email ,
            from_email_smtp_password = from_email_smtp_password ,
        )
    
    # メールアドレスのExcelファイルをS3から取得
    manipulate_s3.s3_file_download(local_upload_path = mail_excel_path)
    time.sleep(2)

    # Excelファイルからメールアドレスをリストで抽出
    receive_mail_list , cc_mail_list , from_email , \
        from_email_smtp_password = mail_list_from_excel(mail_excel_path)
    print(f"変更前 : ")
    print(receive_mail_list)
    receive_mail_list_str = "\n".join(receive_mail_list)
    return render_template(
        "schedule_mail.html" ,
        receive_mail_list_str = receive_mail_list_str ,
        from_email = from_email ,
        from_email_smtp_password = from_email_smtp_password ,
    )




@app.route('/csv_download')
def csv_download():
    directory = os.path.join(app.root_path, 'files') 
    return send_file(os.path.join(directory, output_reins_csv_path), as_attachment=True)

@app.route('/excel_download')
def excel_download():
    directory = os.path.join(app.root_path, 'files') 
    return send_file(os.path.join(directory, output_reins_excel_path), as_attachment=True)



if __name__ == "__main__":
    port_number = 8810
    app.run(port = port_number , debug=True)




