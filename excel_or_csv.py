import openpyxl
import csv


def list_to_csv(to_csv_list: list , csv_path: str = "output.csv"):
    """ 多次元リストのデータをcsvファイルに保存する関数 """
    with open(csv_path, 'w' , encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(to_csv_list)



def csv_to_list(csv_path: str = "output.csv"):
    """ 多次元データを含むcsvからリストに変換 """
    data_list = []
    with open(csv_path, 'r' , encoding="utf-8-sig") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            data_list.append(row)
    return data_list


def list_to_excel(to_excel_list: list , output_excel_path: str = "output.xlsx"):
    workbook = openpyxl.Workbook()
    print("Excelのワークブック起動完了 : workbook = openpyxl.load_workbook()")
    sheet = workbook.active
    print("ワークブックのアクティブ化完了 : sheet = workbook.active")
    
    # 多次元リストのサイズを取得(行ごとで列数に違いがあることを考慮)
    row_num , col_num = len(to_excel_list) , 0
    for row in range(row_num):
        predict_col = len(to_excel_list[row])
        if predict_col > col_num:
            col_num = predict_col
    print(f"row_num : {row_num}")
    print(f"col_num : {col_num}")

    for row in range(row_num):
        for col in range(col_num):
            try:
                print(f"pressed_cell_value : {to_excel_list[row][col]}")
                print(f"row , col : {row} , {col} \n \n")
                sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]
            except IndexError:
                pass
    print("セルの編集可能が証明 : sheet.cell(row=row+1, column=col+1).value = to_excel_list[row][col]")
    workbook.save(output_excel_path)
    

def get_search_option_from_csv(input_csv_path):
    """ 定期実行ツールがcsvファイルから検索方法と条件を取得する関数 """
    search_option_list = csv_to_list(input_csv_path)
    index_of_search_requirement_list = [
        int( search_option_list[1][1] ) ,
        int( search_option_list[2][1] ) ,
    ]
    return index_of_search_requirement_list


def get_search_option_from_excel(input_excel_path):
    """ Excelファイルから多次元リストを抽出する関数 """
    workbook = openpyxl.load_workbook(input_excel_path)
    sheet = workbook.active
    
    index_of_solding_requirement_list = []
    index_of_rental_requirement_list = []
    # 2行目のB列以降のデータを取得してリストに格納
    for cell in sheet[2][1:]:
        if cell == None:
            cell = 0
        index_of_solding_requirement_list.append(cell.value)
    # 3行目のB列以降のデータを取得してリストに格納
    for cell in sheet[3][1:]:
        if cell == None:
            cell = 0
        index_of_rental_requirement_list.append(cell.value)
    return index_of_solding_requirement_list, index_of_rental_requirement_list



def update_search_excel_file(excel_path , search_requirement_index_list):
    """ 検索方法と条件のexcelファイルを編集する関数 """
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row = 1 , column = 1).value = "method"
    sheet.cell(row = 1 , column = 1).value = "requirement"
    sheet.cell(row = 2 , column = 1).value = "select_solding"
    sheet.cell(row = 3 , column = 1).value = "select_rental"

    for loop in range(2):
        index_of_solding_or_rental_list = search_requirement_index_list[loop]
        for small_loop , index_of_solding_or_rental in enumerate(index_of_solding_or_rental_list):
            sheet.cell(row = 2 + loop , column = 2 + small_loop).value = int( index_of_solding_or_rental )
            print(f"row , col : {loop} , {small_loop}")
    
    workbook.save(excel_path)


def update_search_csv_file(csv_path , search_method_name_list , search_requirement_index_list):
    """ 検索方法と条件のcsvファイルを編集する関数 """
    # CSVファイルを読み込みモードで開く
    with open(csv_path, 'r', newline='', encoding='utf-8') as file:
        # CSVファイルを読み込む
        csv_reader = csv.reader(file)
        rows = list(csv_reader)
    # 2行1列目と2行2列目を書き換える
    for loop in range(len(search_method_name_list)):
        rows[1 + loop][0] = search_method_name_list[loop]

        index_of_solding_or_rental_list = search_requirement_index_list[loop]
        for small_loop , index_of_solding_or_rental in enumerate(index_of_solding_or_rental_list):
            rows[1 + loop][1 + small_loop] = str(index_of_solding_or_rental)

    # CSVファイルを書き込みモードで開く
    with open(csv_path, 'w', newline='', encoding='utf-8') as file:
        # 書き込み用のCSVライターを作成
        csv_writer = csv.writer(file)
        # 書き換えたデータをCSVファイルに書き込む
        csv_writer.writerows(rows)


def mail_list_from_excel(mail_excel_path):
    """ Excelファイルからメールのリスト(cc含)を取得する関数 """
    mail_list = []
    workbook = openpyxl.load_workbook(mail_excel_path)
    sheet = workbook.active
    receive_email_number = 100
    for index in range(receive_email_number):
        mail = sheet.cell(row = index + 2 , column = 3).value
        # emailかどうかを判定（「@」「.」の有無）
        if mail is not None:
            if '@' in mail and '.' in mail:
                mail_list.append(mail)
        else:
            break
    # ccのメールのリストを取得
    cc_mail_list = []
    for index in range(len(mail_list)):
        # D列以降を判定
        cc_mail_row_list = []
        for col in range(receive_email_number):
            cc_mail = sheet.cell(row = index + 2 , column = 4 + col).value
            # emailかどうかを判定（「@」「.」の有無）
            if cc_mail is not None:
                if '@' in cc_mail and '.' in cc_mail:
                    cc_mail_row_list.append(cc_mail)
            else:
                break
        cc_mail_list.append(cc_mail_row_list)
    # 送信元メールアドレスとアプリパスワードを取得
    from_email = sheet.cell(row = 2 , column = 1).value
    from_email_smtp_password = sheet.cell(row = 2 , column = 2).value
    return mail_list , cc_mail_list , from_email , from_email_smtp_password


def mail_list_to_excel(receive_mail_list , mail_excel_path):
    """ メールリストのエクセルファイルを変更する関数 """
    workbook = openpyxl.load_workbook(mail_excel_path)
    sheet = workbook.active
    # C列の2行目から順にメールリストの要素を書き込む
    start_row = 2  # 2行目から始める例
    for receive_mail in receive_mail_list:
        if '@' in receive_mail and '.' in receive_mail:
            sheet.cell(row = start_row, column = 3, value = receive_mail)
        start_row += 1
    print(f"len(receive_mail_list) : {len(receive_mail_list)}")
    # メールアドレスが削除されたケース
    loop_count = len(receive_mail_list)
    end_flag = False
    while end_flag == False:
        delete_cell = sheet.cell(row = 2 + loop_count , column = 3).value
        print(f"delete_cell : {delete_cell}")
        if delete_cell != None:
            sheet.cell(row = 2 + loop_count , column = 3).value = None
            print(f"delete_cell : {sheet.cell(row = 2 + loop_count , column = 3).value}")
            loop_count += 1
            print("\n")
        else:
            break
    # 変更を保存
    workbook.save(mail_excel_path)
